from math import pi
from math import sqrt as sq
from math import *

from openpyxl import load_workbook
import xlsxwriter
import openpyxl


def task1():
    print('TASK1')
    print(pi)
    print('pi=%.2f' % pi * 2)


def task2():
    print('\n\nTASK2')
    print(sq(4))


def task3():
    print('\n\nTASK3')
    print(sin(3.14))


def task4_1():
    print('\n\nTASK 4.1 - 4.2')
    wb = load_workbook(filename='lr9.xlsx')
    sheet = wb['Sheet']
    print(sheet['A1'].value)


def task4_3():
    print('\n\nTASK4.3')
    wb = load_workbook(filename='lr9.xlsx')
    wb.create_sheet(title='Sheet', index=0)
    sheet = wb['Sheet']
    sheet['A1'] = 'Valerii was here'
    print(sheet['A1'].value)
    wb.save(filename='lr9.xlsx')


def task4_4():
    print('\n\nTASK4.4')
    wb = xlsxwriter.Workbook('lr9.xlsx', {'strings_to_numbers': True})


def task4_5():
    print('\n\nTASK4.5')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['C1'] = 1
    sheet['C2'] = 2
    sheet['C3'] = 3
    sheet['C4'] = 4
    sheet['C5'] = 5
    sheet['C6'] = 6
    sheet['C7'] = 7
    sheet['C8'] = 8
    sheet['C9'] = 9
    sheet['C10'] = '=SUM(C1:C9)'
    wb.save(filename='lr9.xlsx')


def main():
    task1()
    task2()
    task3()
    task4_1()
    task4_3()
    task4_4()
    task4_5()
